from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from scripts.common.data_paths import (
    EXTRACTED_CURRICULUM_FLOWCHARTS_DIR,
)


YEARS = [
    2022,
    2023,
    2024,
]


# ---------------------------------------------------------
# 경로
# ---------------------------------------------------------

def report_directory() -> Path:
    return (
        EXTRACTED_CURRICULUM_FLOWCHARTS_DIR
        / "comparison"
        / "required_credit_change_report"
    )


def report_path(
    year: int,
) -> Path:
    return (
        report_directory()
        / f"required_credit_change_report_{year}.txt"
    )


def output_path() -> Path:
    return (
        report_directory()
        / "전공필수_학점변화_학과문의용.xlsx"
    )


# ---------------------------------------------------------
# report 파싱
# ---------------------------------------------------------

def read_report(
    year: int,
) -> str:
    path = report_path(year)

    if not path.exists():
        raise FileNotFoundError(
            f"report가 없습니다: {path}\n"
            "먼저 아래 명령을 실행하세요:\n\n"
            "python -m "
            "scripts.curriculum.seed_reconciliation."
            "build_required_credit_change_report --all"
        )

    return path.read_text(
        encoding="utf-8",
    )


def parse_report(
    year: int,
    text: str,
) -> dict[str, object]:
    lines = [
        line.strip()
        for line in text.splitlines()
        if line.strip()
    ]

    if len(lines) < 4:
        raise RuntimeError(
            f"{year}학번 report 형식이 너무 짧습니다."
        )

    total_match = re.fullmatch(
        r"([0-9.]+)학점\s*→\s*([0-9.]+)학점",
        lines[1],
    )

    if not total_match:
        raise RuntimeError(
            f"{year}학번 총학점 행을 읽을 수 없습니다: "
            f"{lines[1]}"
        )

    before_total = float(
        total_match.group(1)
    )

    after_total = float(
        total_match.group(2)
    )

    changes: list[dict[str, object]] = []

    summary = ""
    validation = ""

    for line in lines[2:]:
        if line.startswith("합계:"):
            summary = line
            continue

        if line.startswith("검증:"):
            validation = line
            continue

        if (
            ":" in line
            and "→" in line
        ):
            changes.append(
                parse_change_line(
                    line
                )
            )

    if not validation.startswith(
        "검증: PASS"
    ):
        raise RuntimeError(
            f"{year}학번 report가 PASS 상태가 아닙니다:\n"
            f"{validation}"
        )

    return {
        "year": year,
        "before_total": before_total,
        "after_total": after_total,
        "delta": (
            after_total
            - before_total
        ),
        "changes": changes,
        "summary": summary,
    }


def parse_change_line(
    line: str,
) -> dict[str, object]:
    label, rest = line.split(
        ":",
        1,
    )

    rest = rest.strip()

    impact_match = re.search(
        r"\(([+-]?[0-9.]+)\)$",
        rest,
    )

    if not impact_match:
        raise RuntimeError(
            f"전필 영향 값을 읽을 수 없습니다:\n{line}"
        )

    impact = float(
        impact_match.group(1)
    )

    transition = (
        rest[
            :impact_match.start()
        ]
        .strip()
    )

    if "→" not in transition:
        raise RuntimeError(
            f"변경 전/후를 읽을 수 없습니다:\n{line}"
        )

    before_text, after_text = (
        transition.split(
            "→",
            1,
        )
    )

    return {
        "course": normalize_course_label(
            label.strip()
        ),
        "before": normalize_state_text(
            before_text.strip()
        ),
        "after": normalize_state_text(
            after_text.strip()
        ),
        "impact": impact,
    }


# ---------------------------------------------------------
# 표시 문자열 정리
# ---------------------------------------------------------

def normalize_course_label(
    value: str,
) -> str:
    replacements = {
        "의약품합성학1": "의약품합성학1·2",
    }

    return replacements.get(
        value,
        value,
    )


def normalize_state_text(
    value: str,
) -> str:
    """
    report 내부의
        전필 2
        전선 3
        2학점
    같은 표현을 학과 제출용 표에서는
        전필 2학점
        전선 3학점
    형식으로 통일한다.

    여러 과목이 '+'로 연결된 문장은 그대로 유지한다.
    """

    value = value.strip()

    if value == "신규":
        return value

    # "전필 2" / "전선 3" -> "전필 2학점" / "전선 3학점"
    value = re.sub(
        r"\b(전필|전선)\s+([0-9]+(?:\.[0-9]+)?)\b(?!학점)",
        r"\1 \2학점",
        value,
    )

    # 변경 후가 단순 "2학점"처럼 되어 있으면
    # 변경 전의 이수구분을 여기에서 추론하지 않는다.
    # report가 명시한 내용만 유지한다.

    return value


def format_number(
    value: float,
) -> str:
    if value.is_integer():
        return str(
            int(value)
        )

    return f"{value:g}"


def format_delta(
    value: float,
    unit: bool = True,
) -> str:
    if value > 0:
        text = (
            f"+{format_number(value)}"
        )

    else:
        text = format_number(
            value
        )

    if unit:
        return (
            f"{text}학점"
        )

    return text


# ---------------------------------------------------------
# 스타일
# ---------------------------------------------------------

TITLE_FILL = PatternFill(
    "solid",
    fgColor="1F4E78",
)

SUMMARY_HEADER_FILL = PatternFill(
    "solid",
    fgColor="5B9BD5",
)

SECTION_FILL = PatternFill(
    "solid",
    fgColor="D9EAF7",
)

TABLE_HEADER_FILL = PatternFill(
    "solid",
    fgColor="E2F0D9",
)

POSITIVE_FILL = PatternFill(
    "solid",
    fgColor="E2F0D9",
)

NEGATIVE_FILL = PatternFill(
    "solid",
    fgColor="FCE4D6",
)

TOTAL_FILL = PatternFill(
    "solid",
    fgColor="F2F2F2",
)

WHITE_FONT = Font(
    color="FFFFFF",
    bold=True,
)

THIN_SIDE = Side(
    style="thin",
    color="BFBFBF",
)

THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)


def apply_border(
    ws,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        for cell in row:
            cell.border = THIN_BORDER


# ---------------------------------------------------------
# workbook 작성
# ---------------------------------------------------------

def build_workbook(
    reports: list[dict[str, object]],
) -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "전필 학점 변화"

    # -------------------------------------------------
    # 전체 제목
    # -------------------------------------------------

    ws.merge_cells(
        "A1:E1"
    )

    title_cell = ws["A1"]

    title_cell.value = (
        "학번별 전공필수 학점 변화"
    )

    title_cell.fill = TITLE_FILL

    title_cell.font = Font(
        color="FFFFFF",
        bold=True,
        size=16,
    )

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    ws.row_dimensions[1].height = 30

    # -------------------------------------------------
    # 요약
    # -------------------------------------------------

    summary_header_row = 3

    headers = [
        "학번",
        "기존 전필",
        "변경 후",
        "순변화",
    ]

    for col, value in enumerate(
        headers,
        start=1,
    ):
        cell = ws.cell(
            row=summary_header_row,
            column=col,
            value=value,
        )

        cell.fill = SUMMARY_HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for index, report in enumerate(
        reports,
        start=1,
    ):
        row = (
            summary_header_row
            + index
        )

        year = int(
            report["year"]
        )

        before_total = float(
            report["before_total"]
        )

        after_total = float(
            report["after_total"]
        )

        delta = float(
            report["delta"]
        )

        values = [
            f"{year}학번",
            (
                f"{format_number(before_total)}학점"
            ),
            (
                f"{format_number(after_total)}학점"
            ),
            format_delta(
                delta
            ),
        ]

        for col, value in enumerate(
            values,
            start=1,
        ):
            cell = ws.cell(
                row=row,
                column=col,
                value=value,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        delta_cell = ws.cell(
            row=row,
            column=4,
        )

        delta_cell.font = Font(
            bold=True,
        )

        if delta > 0:
            delta_cell.fill = POSITIVE_FILL

        elif delta < 0:
            delta_cell.fill = NEGATIVE_FILL

    summary_end_row = (
        summary_header_row
        + len(reports)
    )

    apply_border(
        ws,
        summary_header_row,
        summary_end_row,
        1,
        4,
    )

    # -------------------------------------------------
    # 학번별 상세
    # -------------------------------------------------

    row = (
        summary_end_row
        + 3
    )

    for report in reports:
        year = int(
            report["year"]
        )

        before_total = float(
            report["before_total"]
        )

        after_total = float(
            report["after_total"]
        )

        delta = float(
            report["delta"]
        )

        changes = list(
            report["changes"]
        )

        # ---------------------------------------------
        # 학번 섹션 제목
        # ---------------------------------------------

        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=5,
        )

        section_cell = ws.cell(
            row=row,
            column=1,
        )

        section_cell.value = (
            f"{year}학번"
        )

        section_cell.fill = SECTION_FILL

        section_cell.font = Font(
            bold=True,
            size=13,
        )

        section_cell.alignment = Alignment(
            vertical="center",
        )

        ws.row_dimensions[row].height = 24

        row += 1

        # ---------------------------------------------
        # 118 → 122 식 총합 표시
        # ---------------------------------------------

        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=5,
        )

        total_change_cell = ws.cell(
            row=row,
            column=1,
        )

        total_change_cell.value = (
            f"{format_number(before_total)}학점"
            f"  →  "
            f"{format_number(after_total)}학점"
            f"   "
            f"({format_delta(delta)})"
        )

        total_change_cell.font = Font(
            bold=True,
            size=15,
        )

        total_change_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        ws.row_dimensions[row].height = 28

        row += 2

        # ---------------------------------------------
        # 상세 표 헤더
        # ---------------------------------------------

        detail_header_row = row

        detail_headers = [
            "변경 과목",
            "변경 전",
            "→",
            "변경 후",
            "전필 영향",
        ]

        for col, value in enumerate(
            detail_headers,
            start=1,
        ):
            cell = ws.cell(
                row=row,
                column=col,
                value=value,
            )

            cell.fill = TABLE_HEADER_FILL

            cell.font = Font(
                bold=True,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        row += 1

        detail_start_row = row

        # ---------------------------------------------
        # 변화 항목
        # ---------------------------------------------

        for change in changes:
            impact = float(
                change["impact"]
            )

            values = [
                str(
                    change["course"]
                ),
                str(
                    change["before"]
                ),
                "→",
                str(
                    change["after"]
                ),
                format_delta(
                    impact
                ),
            ]

            for col, value in enumerate(
                values,
                start=1,
            ):
                cell = ws.cell(
                    row=row,
                    column=col,
                    value=value,
                )

                cell.alignment = Alignment(
                    horizontal=(
                        "center"
                        if col in [
                            3,
                            5,
                        ]
                        else "left"
                    ),
                    vertical="center",
                    wrap_text=True,
                )

            impact_cell = ws.cell(
                row=row,
                column=5,
            )

            impact_cell.font = Font(
                bold=True,
            )

            if impact > 0:
                impact_cell.fill = (
                    POSITIVE_FILL
                )

            elif impact < 0:
                impact_cell.fill = (
                    NEGATIVE_FILL
                )

            # 긴 과목명/변경 설명을 고려
            ws.row_dimensions[
                row
            ].height = 36

            row += 1

        detail_end_row = (
            row - 1
        )

        apply_border(
            ws,
            detail_header_row,
            detail_end_row,
            1,
            5,
        )

        # ---------------------------------------------
        # 합계
        # ---------------------------------------------

        row += 1

        ws.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=5,
        )

        total_cell = ws.cell(
            row=row,
            column=1,
        )

        total_cell.value = (
            "합계: "
            f"{format_number(before_total)}학점 "
            f"{format_delta(delta)} "
            f"= "
            f"{format_number(after_total)}학점"
        )

        total_cell.fill = TOTAL_FILL

        total_cell.font = Font(
            bold=True,
        )

        total_cell.alignment = Alignment(
            horizontal="right",
            vertical="center",
        )

        ws.row_dimensions[row].height = 24

        row += 3

    # -------------------------------------------------
    # 주석
    # -------------------------------------------------

    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row + 1,
        end_column=5,
    )

    note_cell = ws.cell(
        row=row,
        column=1,
    )

    note_cell.value = (
        "※ 본 자료는 학번별 교육과정 baseline과 현재 정리된 "
        "curriculum seed를 비교하여 전공필수 과목 학점 변화를 "
        "산정한 확인용 자료입니다. 공식 졸업요건의 최종 적용 여부는 "
        "학과 확인이 필요합니다."
    )

    note_cell.font = Font(
        size=9,
        color="666666",
    )

    note_cell.alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )

    # -------------------------------------------------
    # 열 너비
    # -------------------------------------------------

    column_widths = {
        "A": 31,
        "B": 42,
        "C": 6,
        "D": 53,
        "E": 14,
    }

    for column, width in (
        column_widths.items()
    ):
        ws.column_dimensions[
            column
        ].width = width

    # -------------------------------------------------
    # 인쇄 설정
    # -------------------------------------------------

    ws.freeze_panes = "A3"

    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = (
        "landscape"
    )

    ws.page_setup.paperSize = (
        ws.PAPERSIZE_A4
    )

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = (
        True
    )

    ws.print_options.horizontalCentered = (
        True
    )

    ws.print_area = (
        f"A1:E{row + 1}"
    )

    ws.oddFooter.center.text = (
        "학번별 전공필수 학점 변화"
    )

    ws.oddFooter.right.text = (
        "Page &P / &N"
    )

    return wb


# ---------------------------------------------------------
# main
# ---------------------------------------------------------

def main() -> None:
    reports = []

    for year in YEARS:
        reports.append(
            parse_report(
                year,
                read_report(
                    year
                ),
            )
        )

    wb = build_workbook(
        reports
    )

    target = output_path()

    target.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    wb.save(
        target
    )

    print(
        "학과 문의용 Excel 생성 완료"
    )

    print(
        f"output: {target}"
    )


if __name__ == "__main__":
    main()